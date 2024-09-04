from typing import Iterable

import re
from pathlib import Path

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from finparse.models import Card, Transaction, Currency

title_pattern = re.compile(r"לכרטיס\s(.*?)\sהמסתיים.*(\d{4})$")
currency_pattern = re.compile(r"\[\$(.*?)\]")


def get_currency(number_formatting: str) -> Currency:
    match = currency_pattern.search(number_formatting)
    return Currency(match.group(1))


def parse_workbook(workbook_path: Path) -> Iterable[Card]:
    workbook: Workbook = openpyxl.load_workbook(
        workbook_path, read_only=True, data_only=True, keep_links=False
    )

    sheet: Worksheet = workbook[workbook.sheetnames[0]]
    cell: Cell = sheet["A1"]

    match = title_pattern.search(cell.value)
    card_name = match.group(1)
    last_four_digits = match.group(2)

    card = Card(name=card_name, last_4_digits=last_four_digits)

    row_idx = 2
    # Start with 2nd row, and iterate until we find the header of the transactions
    while not str(sheet.cell(row_idx, 1).value).endswith("עסקה"):
        row_idx += 1

    # Next row_idx is the first transaction
    row_idx += 1

    row: tuple[Cell, ...]
    for row in sheet.iter_rows(row_idx):
        if not row[0].value:
            break

        date, description, foreign_cost, local_cost, _, category, notes = row
        amount = str(local_cost.value)
        currency = get_currency(local_cost.number_format)
        foreign_amount = str(foreign_cost.value)
        foreign_currency = get_currency(foreign_cost.number_format)

        card.transactions.append(
            Transaction(
                date=date.value,
                description=description.value,
                amount=amount,
                currency=currency,
                foreign_amount=foreign_amount,
                foreign_currency=foreign_currency,
                category=category.value,
                notes=notes.value,
            )
        )
        row_idx += 1

    yield card
